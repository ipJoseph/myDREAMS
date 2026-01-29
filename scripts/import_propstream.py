#!/usr/bin/env python3
"""
PropStream Excel Importer

Imports property data from PropStream Excel exports into the normalized schema.
Creates parcels and listings (for MLS-active properties).

Usage:
    python scripts/import_propstream.py /path/to/export.xlsx

DEV ONLY - do not run in production without review.
"""

import hashlib
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).parent.parent
DB_PATH = PROJECT_ROOT / 'data' / 'dreams.db'

# County name normalization
COUNTY_NORMALIZE = {
    'cherokee': 'Cherokee',
    'cherokee county': 'Cherokee',
    'clay': 'Clay',
    'clay county': 'Clay',
    'graham': 'Graham',
    'graham county': 'Graham',
    'haywood': 'Haywood',
    'haywood county': 'Haywood',
    'henderson': 'Henderson',
    'henderson county': 'Henderson',
    'jackson': 'Jackson',
    'jackson county': 'Jackson',
    'macon': 'Macon',
    'macon county': 'Macon',
    'madison': 'Madison',
    'madison county': 'Madison',
    'swain': 'Swain',
    'swain county': 'Swain',
    'transylvania': 'Transylvania',
    'transylvania county': 'Transylvania',
    'buncombe': 'Buncombe',
    'buncombe county': 'Buncombe',
}


def generate_id(prefix: str, *args) -> str:
    """Generate a deterministic ID from input values."""
    data = '|'.join(str(a) for a in args if a)
    hash_val = hashlib.md5(data.encode()).hexdigest()[:12]
    return f"{prefix}_{hash_val}"


def normalize_county(county: str) -> str:
    """Normalize county name."""
    if not county:
        return 'Unknown'
    return COUNTY_NORMALIZE.get(county.lower().strip(), county.strip())


def normalize_address(address: str, unit: str = None) -> str:
    """Normalize address for matching."""
    if not address:
        return ''

    addr = address.strip()

    # Add unit if provided
    if unit:
        addr = f"{addr} #{unit}"

    # Standardize common abbreviations (basic - could expand)
    replacements = [
        (r'\bSt\b', 'St'),
        (r'\bStreet\b', 'St'),
        (r'\bRd\b', 'Rd'),
        (r'\bRoad\b', 'Rd'),
        (r'\bDr\b', 'Dr'),
        (r'\bDrive\b', 'Dr'),
        (r'\bLn\b', 'Ln'),
        (r'\bLane\b', 'Ln'),
        (r'\bTrl\b', 'Trl'),
        (r'\bTrail\b', 'Trl'),
        (r'\bCir\b', 'Cir'),
        (r'\bCircle\b', 'Cir'),
        (r'\bAve\b', 'Ave'),
        (r'\bAvenue\b', 'Ave'),
        (r'\bCt\b', 'Ct'),
        (r'\bCourt\b', 'Ct'),
        (r'\bPl\b', 'Pl'),
        (r'\bPlace\b', 'Pl'),
        (r'\bHwy\b', 'Hwy'),
        (r'\bHighway\b', 'Hwy'),
    ]

    for pattern, replacement in replacements:
        addr = re.sub(pattern, replacement, addr, flags=re.IGNORECASE)

    return addr


def safe_int(value) -> int:
    """Safely convert to int."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def safe_float(value) -> float:
    """Safely convert to float."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_date(value) -> str:
    """Safely convert to ISO date string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        return value[:10] if len(value) >= 10 else value
    return None


def import_propstream(xlsx_path: str):
    """Import PropStream Excel export."""

    print("=" * 60)
    print("PROPSTREAM IMPORT")
    print("=" * 60)
    print(f"Source: {xlsx_path}")
    print(f"Database: {DB_PATH}")
    print()

    # Load Excel
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb['Properties']

    # Get headers
    headers = [cell.value for cell in sheet[1]]
    print(f"Found {len(headers)} columns, {sheet.max_row - 1} rows")

    # Create column index
    col_idx = {h: i for i, h in enumerate(headers)}

    # Connect to DB
    conn = sqlite3.connect(DB_PATH)

    parcels_created = 0
    parcels_updated = 0
    listings_created = 0
    errors = 0

    print("\nImporting...")

    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Extract values using column index
            def get(col_name, default=None):
                idx = col_idx.get(col_name)
                if idx is None:
                    return default
                val = row[idx]
                return val if val is not None else default

            # === PARCEL DATA ===
            apn = str(get('APN', '')).strip() or None
            county = normalize_county(get('County'))
            address_raw = get('Address', '')
            unit = get('Unit #')
            address = normalize_address(address_raw, unit)
            city = get('City', '')
            state = get('State', 'NC')
            zip_code = str(get('Zip', ''))[:10] if get('Zip') else None

            # Generate parcel ID
            if apn and county:
                parcel_id = generate_id('prc', apn, county)
            elif address and city:
                parcel_id = generate_id('prc', address.lower(), city.lower())
            else:
                parcel_id = generate_id('prc', str(row_num))

            # Owner info
            owner1_first = get('Owner 1 First Name', '')
            owner1_last = get('Owner 1 Last Name', '')
            owner2_first = get('Owner 2 First Name', '')
            owner2_last = get('Owner 2 Last Name', '')
            owner_name = f"{owner1_first} {owner1_last}".strip() or None
            owner_name_2 = f"{owner2_first} {owner2_last}".strip() or None

            # Check if parcel exists
            existing = conn.execute(
                "SELECT id FROM parcels WHERE id = ?", (parcel_id,)
            ).fetchone()

            if existing:
                # Update existing parcel
                conn.execute("""
                    UPDATE parcels SET
                        apn = COALESCE(?, apn),
                        county = COALESCE(?, county),
                        address = COALESCE(?, address),
                        address_raw = COALESCE(?, address_raw),
                        city = COALESCE(?, city),
                        zip = COALESCE(?, zip),
                        owner_name = COALESCE(?, owner_name),
                        owner_name_2 = COALESCE(?, owner_name_2),
                        owner_occupied = COALESCE(?, owner_occupied),
                        owner_phone = COALESCE(?, owner_phone),
                        owner_email = COALESCE(?, owner_email),
                        mailing_address = COALESCE(?, mailing_address),
                        mailing_city = COALESCE(?, mailing_city),
                        mailing_state = COALESCE(?, mailing_state),
                        mailing_zip = COALESCE(?, mailing_zip),
                        acreage = COALESCE(?, acreage),
                        land_use = COALESCE(?, land_use),
                        assessed_value = COALESCE(?, assessed_value),
                        last_sale_date = COALESCE(?, last_sale_date),
                        last_sale_amount = COALESCE(?, last_sale_amount),
                        updated_at = ?
                    WHERE id = ?
                """, (
                    apn, county, address, address_raw, city, zip_code,
                    owner_name, owner_name_2,
                    get('Owner Occupied'),
                    get('Mobile') or get('Landline'),
                    get('Email'),
                    get('Mailing Address'),
                    get('Mailing City'),
                    get('Mailing State'),
                    str(get('Mailing Zip', ''))[:10] if get('Mailing Zip') else None,
                    safe_float(get('Lot Size Sqft')) / 43560 if get('Lot Size Sqft') else None,  # Sqft to acres
                    get('Property Type'),
                    safe_int(get('Total Assessed Value')),
                    safe_date(get('Last Sale Date')),
                    safe_int(get('Last Sale Amount')),
                    datetime.now().isoformat(),
                    parcel_id
                ))
                parcels_updated += 1
            else:
                # Create new parcel
                conn.execute("""
                    INSERT INTO parcels (
                        id, apn, county, state,
                        address, address_raw, city, zip,
                        acreage, land_use,
                        owner_name, owner_name_2, owner_occupied,
                        owner_phone, owner_email,
                        mailing_address, mailing_city, mailing_state, mailing_zip,
                        assessed_value, last_sale_date, last_sale_amount
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    parcel_id, apn, county, state,
                    address, address_raw, city, zip_code,
                    safe_float(get('Lot Size Sqft')) / 43560 if get('Lot Size Sqft') else None,
                    get('Property Type'),
                    owner_name, owner_name_2, get('Owner Occupied'),
                    get('Mobile') or get('Landline'),
                    get('Email'),
                    get('Mailing Address'),
                    get('Mailing City'),
                    get('Mailing State'),
                    str(get('Mailing Zip', ''))[:10] if get('Mailing Zip') else None,
                    safe_int(get('Total Assessed Value')),
                    safe_date(get('Last Sale Date')),
                    safe_int(get('Last Sale Amount'))
                ))
                parcels_created += 1

            # === LISTING DATA (if MLS active) ===
            mls_status = get('MLS Status')
            mls_amount = safe_int(get('MLS Amount'))

            if mls_status and mls_amount:
                # Has MLS data - create listing
                # Use APN + status + amount as unique identifier (no MLS# in PropStream)
                listing_id = generate_id('lst', apn or address, mls_status, mls_amount)

                conn.execute("""
                    INSERT OR REPLACE INTO listings (
                        id, parcel_id,
                        mls_source, mls_number,
                        status, list_price, list_date,
                        beds, baths, sqft, year_built, property_type,
                        hoa_fee,
                        listing_agent_name, listing_agent_phone, listing_agent_email,
                        listing_office_name, listing_office_id,
                        source, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    listing_id, parcel_id,
                    'PropStream', None,  # No MLS# from PropStream
                    mls_status, mls_amount,
                    safe_date(get('MLS Date')),
                    safe_int(get('Bedrooms')),
                    safe_float(get('Total Bathrooms')),
                    safe_int(get('Building Sqft')),
                    safe_int(get('Year Built')),
                    get('Property Type'),
                    None,  # HOA not in PropStream
                    get('MLS Agent Name'),
                    get('MLS Agent Phone'),
                    get('MLS Agent E-Mail'),
                    get('MLS Brokerage Name'),
                    get('MLS Brokerage Phone'),
                    'propstream_11county',
                    datetime.now().isoformat()
                ))
                listings_created += 1

            if row_num % 1000 == 0:
                print(f"  Processed {row_num - 1} rows...")
                conn.commit()

        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  Error row {row_num}: {e}")
            continue

    conn.commit()
    conn.close()

    print()
    print("=" * 60)
    print("IMPORT COMPLETE")
    print("=" * 60)
    print(f"Parcels created: {parcels_created}")
    print(f"Parcels updated: {parcels_updated}")
    print(f"Listings created: {listings_created}")
    print(f"Errors: {errors}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_propstream.py <xlsx_path>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not Path(xlsx_path).exists():
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    import_propstream(xlsx_path)


if __name__ == '__main__':
    main()
